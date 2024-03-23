from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_page
from django.http import HttpResponse
from django.db.models import Q

import openpyxl
from io import BytesIO

from bio.models import Bio
from link.models import Link
from publication.models import Publication
from settings.models import Settings

from search.query import text

# Create your views here.

# @cache_page(0)
def search(request):

    query = request.GET.get('query')

    if query != '' and query != None:
        bios = Bio.objects.filter(text(query))
    else:
        return redirect('homepage')

    response = {
        'query': query,
        'bios': bios,
    }

    if not request.user.is_anonymous:
        response['settings'] = get_object_or_404(Settings, user=request.user)

        # footer status informations
        response['status'] = {'bios':Bio.objects.count(),
                              'links' : Link.objects.count(),
                              'publications' : Publication.objects.count(),
        }

    return render(request, 'search.html', response)

def group(request):

    if request.LANGUAGE_CODE == 'en':
        college = request.GET.get('college')
        academic_title = request.GET.get('academic_title')

        if college == 'All' and academic_title == 'All':
            bios = Bio.objects.all()
        elif college == 'All':
            bios = Bio.objects.filter(Q(academic_title=academic_title))
        elif academic_title == 'All':
            bios = Bio.objects.filter(Q(college=college))
        else:
            bios = Bio.objects.filter(Q(college=college) & Q(academic_title=academic_title))

        response = {
            'college': college,
            'academic_title': academic_title,
            'bios': bios,
        }

    else:
        ar_college = request.GET.get('ar_college')
        ar_academic_title = request.GET.get('ar_academic_title')

        if ar_college == 'All' and ar_academic_title == 'All':
            bios = Bio.objects.all()
        elif ar_college == 'All':
            bios = Bio.objects.filter(Q(ar_academic_title=ar_academic_title))
        elif ar_academic_title == 'All':
            bios = Bio.objects.filter(Q(ar_college=ar_college))
        else:
            bios = Bio.objects.filter(Q(ar_college=ar_college) & Q(ar_academic_title=ar_academic_title))

        response = {
            'ar_college': ar_college,
            'ar_academic_title': ar_academic_title,
            'bios': bios,
        }

    if not request.user.is_anonymous:
        response['settings'] = get_object_or_404(Settings, user=request.user)

        # footer status informations
        response['status'] = {'bios':Bio.objects.count(),
                              'links' : Link.objects.count(),
                              'publications' : Publication.objects.count(),
        }

    return render(request, 'search.html', response)

@login_required
def excel(request):

    if not request.user.is_superuser:
        return redirect('homepage')

    if request.LANGUAGE_CODE == 'en':
        college = request.GET.get('college')
        academic_title = request.GET.get('academic_title')
    else:
        college = request.GET.get('ar_college')
        academic_title = request.GET.get('ar_academic_title')

    if college == 'All' and academic_title == 'All':
        bios = Bio.objects.all()
    elif college == 'All':
        bios = Bio.objects.filter(Q(academic_title=academic_title))
    elif academic_title == 'All':
        bios = Bio.objects.filter(Q(college=college))
    else:
        bios = Bio.objects.filter(Q(college=college) & Q(academic_title=academic_title))


    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row=1, column=2).value = 'ar_full_name'
    sheet.cell(row=1, column=3).value = 'full_name'
    sheet.cell(row=1, column=4).value = 'ar_nationality'
    sheet.cell(row=1, column=5).value = 'ar_academic_title'
    sheet.cell(row=1, column=6).value = 'ar_college'
    sheet.cell(row=1, column=7).value = 'ar_department'
    sheet.cell(row=1, column=8).value = 'ar_major'
    sheet.cell(row=1, column=9).value = 'ar_specialty'
    sheet.cell(row=1, column=10).value = 'ar_position'
    sheet.cell(row=1, column=11).value = 'ar_occupation'
    sheet.cell(row=1, column=12).value = 'ar_mother_lang'
    sheet.cell(row=1, column=13).value = 'ar_other_langs'
    sheet.cell(row=1, column=14).value = 'ar_state'
    sheet.cell(row=1, column=15).value = 'ar_district'
    sheet.cell(row=1, column=16).value = 'ar_hiring_date'
    sheet.cell(row=1, column=17).value = 'ar_direct_date'
    sheet.cell(row=1, column=18).value = 'surname'
    sheet.cell(row=1, column=19).value = 'birthday'
    sheet.cell(row=1, column=20).value = 'nationality'
    sheet.cell(row=1, column=21).value = 'gender'
    sheet.cell(row=1, column=22).value = 'degree'
    sheet.cell(row=1, column=23).value = 'academic_title'
    sheet.cell(row=1, column=24).value = 'college'
    sheet.cell(row=1, column=25).value = 'department'
    sheet.cell(row=1, column=26).value = 'major'
    sheet.cell(row=1, column=27).value = 'specialty'
    sheet.cell(row=1, column=28).value = 'position'
    sheet.cell(row=1, column=29).value = 'occupation'
    sheet.cell(row=1, column=30).value = 'mother_lang'
    sheet.cell(row=1, column=31).value = 'other_langs'
    sheet.cell(row=1, column=32).value = 'personal_email'
    sheet.cell(row=1, column=33).value = 'work_email'
    sheet.cell(row=1, column=34).value = 'create_date'
    sheet.cell(row=1, column=35).value = 'update_date'
        
    for i, bio in enumerate(bios, 2):

        sheet.cell(row=i, column=2).value = bio.ar_full_name 
        sheet.cell(row=i, column=3).value = bio.full_name
        sheet.cell(row=i, column=4).value = bio.ar_nationality
        sheet.cell(row=i, column=5).value = bio.ar_academic_title
        sheet.cell(row=i, column=6).value = bio.ar_college
        sheet.cell(row=i, column=7).value = bio.ar_department
        sheet.cell(row=i, column=8).value = bio.ar_major
        sheet.cell(row=i, column=9).value = bio.ar_specialty
        sheet.cell(row=i, column=10).value = bio.ar_position
        sheet.cell(row=i, column=11).value = bio.ar_occupation
        sheet.cell(row=i, column=12).value = bio.ar_mother_lang
        sheet.cell(row=i, column=13).value = bio.ar_other_langs
        sheet.cell(row=i, column=14).value = bio.ar_state
        sheet.cell(row=i, column=15).value = bio.ar_district
        sheet.cell(row=i, column=16).value = str(bio.ar_hiring_date)
        sheet.cell(row=i, column=17).value = str(bio.ar_direct_date)
        sheet.cell(row=i, column=18).value = bio.surname
        sheet.cell(row=i, column=19).value = str(bio.birthday)
        sheet.cell(row=i, column=20).value = bio.nationality
        sheet.cell(row=i, column=21).value = bio.gender
        sheet.cell(row=i, column=22).value = bio.degree
        sheet.cell(row=i, column=23).value = bio.academic_title
        sheet.cell(row=i, column=24).value = bio.college
        sheet.cell(row=i, column=25).value = bio.department
        sheet.cell(row=i, column=26).value = bio.major
        sheet.cell(row=i, column=27).value = bio.specialty
        sheet.cell(row=i, column=28).value = bio.position
        sheet.cell(row=i, column=29).value = bio.occupation
        sheet.cell(row=i, column=30).value = bio.mother_lang
        sheet.cell(row=i, column=31).value = bio.other_langs
        sheet.cell(row=i, column=32).value = bio.personal_email
        sheet.cell(row=i, column=33).value = bio.work_email
        sheet.cell(row=i, column=34).value = str(bio.create_date)
        sheet.cell(row=i, column=35).value = str(bio.update_date)

    buffer = BytesIO()
    wb.save(buffer)

    buffer.seek(0)

    # create response
    response = HttpResponse(content_type=f'application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="qu-staff-cv-data.xlsx"'
    response.write(buffer.read())

    return response

